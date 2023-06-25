import os
import traceback

from PyQt5.QtCore import Qt, QSize, QThread, QModelIndex, QAbstractItemModel, QRegularExpression
from PyQt5.QtGui import QDoubleValidator, QRegularExpressionValidator
from PyQt5.QtWidgets import QApplication, QStyleFactory, QDialog, QDialogButtonBox, QWidget, QFileDialog, \
    QAbstractButton, QMessageBox, QTableWidget, QComboBox, QStyledItemDelegate, QTableWidgetItem, QHeaderView
from PyQt5 import QtCore, QtGui
import sys

from openpyxl import load_workbook

from UI.UIModelCal import Ui_frmModelCal
from UICore.SpModel import modelCal
from UICore.log4p import Log
from UICore.DataFactory import workspaceFactory
from osgeo import ogr
from UICore.Gv import DataType

Slot = QtCore.pyqtSlot

log = Log(__name__)

Grid_neccessary_fields = {
    '标准单元编号': ['UnitID', 'num'],
    '未来就业岗位': ['PlaJOB', 'num'],
    # '规划居住建筑总面积': ['PlaBS', 'num'],
    '单元现状人口总数': ['CurPOP', 'num'],
    '单元现状就业岗位总数': ['CurJOB', 'num'],
    '单元类型': ['UnitType', 'str']
}

Potential_Land_neccessary_fields = {
    '居住地块编号': ['LandID', 'num'],
    '居住地块用地路径类型': ['Type', 'num'],
    '新建居住建筑潜力面积': ['R_Po', 'num'],
    '是否在地铁站范围内': ['Metro_IF', 'num'],
    '可享用的公服面积': ['PubService', 'num']
}

config_path = r"config/params.xlsx"

class frmModelCal(QWidget, Ui_frmModelCal):
    def __init__(self, parent=None):
        super(frmModelCal, self).__init__(parent=parent)
        self.parent = parent
        self.setupUi(self)

        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(9)
        self.buttonBox.setStandardButtons(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.buttonBox.button(QDialogButtonBox.Ok).setFont(font)
        self.buttonBox.button(QDialogButtonBox.Ok).setText("运行")
        self.buttonBox.button(QDialogButtonBox.Cancel).setFont(font)
        self.buttonBox.button(QDialogButtonBox.Cancel).setText("取消")

        self.btn_addGridFile.clicked.connect(self.addGridFile_clicked)
        self.btn_addPotentialLandFile.clicked.connect(self.addPotentialLandFile_clicked)
        self.buttonBox.clicked.connect(self.buttonBox_clicked)

        self.validateValue()
        self.bFirstShow = True

    #  初始化模型参数
    def init_model_param(self):
        wb = None
        try:
            if os.path.exists(config_path):
                wb = load_workbook(config_path, read_only=True)
                lst_names = wb.sheetnames

                if "Potential_Constraint" not in lst_names or "IndicatorWeight" not in lst_names:
                    raise Exception("error")

                sheet = wb["Potential_Constraint"]
                self.txt_type_csgx.setText(str(sheet.cell(2, 3).value))
                self.txt_type_tdzb.setText(str(sheet.cell(3, 3).value))
                self.txt_type_zzgjz.setText(str(sheet.cell(4, 3).value))
                self.txt_type_gygjz.setText(str(sheet.cell(5, 3).value))

                sheet = wb["IndicatorWeight"]
                self.txt_indicator_xzjjl.setText(str(sheet.cell(2, 2).value))
                self.txt_indicator_ccjzl.setText(str(sheet.cell(3, 2).value))
                self.txt_indicator_xtzs.setText(str(sheet.cell(4, 2).value))
                self.txt_indicator_zzphzs.setText(str(sheet.cell(5, 2).value))
                self.txt_indicator_jtkdx.setText(str(sheet.cell(6, 2).value))
                self.txt_indicator_ggfwsp.setText(str(sheet.cell(7, 2).value))
            else:
                raise Exception("error")
        except:
            log.warning("读取参数配置文件{}发生错误，使用默认参数".format(config_path))

            self.txt_type_csgx.setText("1")
            self.txt_type_tdzb.setText("1")
            self.txt_type_zzgjz.setText("1")
            self.txt_type_gygjz.setText("1")

            self.txt_indicator_xzjjl.setText("0.5")
            self.txt_indicator_xtzs.setText("0.1")
            self.txt_indicator_ccjzl.setText("0.1")
            self.txt_indicator_jtkdx.setText("0.1")
            self.txt_indicator_ggfwsp.setText("0.1")
            self.txt_indicator_zzphzs.setText("0.1")
        finally:
            if wb is not None:
                wb.close()

    def showEvent(self, QShowEvent):
        self.table_init(self.tbl_GridField)
        self.table_init(self.tbl_PotentialLandField)

        if self.bFirstShow:
            self.init_model_param()
            self.bFirstShow = False

    def resizeEvent(self, QResizeEvent):
        self.table_init(self.tbl_GridField)
        self.table_init(self.tbl_PotentialLandField)

    def table_init(self, tbl: QTableWidget):
        tbl.setColumnCount(2)
        tbl.setHorizontalHeaderLabels(["说明", "字段"])
        tbl.horizontalHeader().setSectionsMovable(False)
        # tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        tbl.setColumnWidth(0, tbl.width() * 0.45)
        tbl.setColumnWidth(1, tbl.width() * 0.5)

    def sizeHint(self):
        return QSize(self.parent.width() * 0.3, self.parent.height())

    # 标准单元图层
    def addGridFile_clicked(self):
        try:
            status, fileName, datasource = self.add_spatial_data()

            if not status:
                return

            bValidate_name, bValidate_type = self.add_neccesary_field_to_item(datasource, self.tbl_GridField, Grid_neccessary_fields)
            if not bValidate_name and bValidate_type:
                log.error("标准单元图层缺失必要字段，请补齐！", dialog=True)
            elif not bValidate_type and bValidate_name:
                log.error("标准单元图层字段类型不匹配，请检查！", dialog=True)
            elif not bValidate_type and not bValidate_name:
                log.error("标准单元图层字段缺失必要字段，并且存在字段类型不匹配，请检查！", dialog=True)
            self.txt_GridFile.setText(fileName)
        except:
            log.error(traceback.format_exc())

    # 潜力用地图层
    def addPotentialLandFile_clicked(self):
        try:
            status, fileName, datasource = self.add_spatial_data()
            if not status:
                return

            bValidate_name, bValidate_type = self.add_neccesary_field_to_item(datasource, self.tbl_PotentialLandField, Potential_Land_neccessary_fields)
            if not bValidate_name and bValidate_type:
                log.error("潜力用地图层缺失必要字段，请补齐！", dialog=True)
            elif not bValidate_type and bValidate_name:
                log.error("潜力用地图层字段类型不匹配，请检查！", dialog=True)
            elif not bValidate_type and not bValidate_name:
                log.error("潜力用地图层字段缺失必要字段，并且存在字段类型不匹配，请检查！", dialog=True)

            self.txt_PotentialLandFile.setText(fileName)
        except:
            log.error(traceback.format_exc())

    # 将必要字段添加到列表
    def add_neccesary_field_to_item(self, datasource, tbl, neccessary_fileds):
        field_num_list = []
        field_str_list = []
        field_num_list.append("")
        field_str_list.append("")
        layer = datasource.GetLayer(0)
        layerDefn = layer.GetLayerDefn()
        for i in range(layerDefn.GetFieldCount()):
            fieldName = layerDefn.GetFieldDefn(i).GetName()
            if layerDefn.GetFieldDefn(i).GetType() == ogr.OFTInteger or layerDefn.GetFieldDefn(i).GetType() == \
                    ogr.OFTInteger64 or layerDefn.GetFieldDefn(i).GetType() == ogr.OFTReal:
                field_num_list.append(fieldName)
            elif layerDefn.GetFieldDefn(i).GetType() == ogr.OFTString or layerDefn.GetFieldDefn(i).GetType() == \
                ogr.OFTWideString:
                field_str_list.append(fieldName)

        bValidate_name = True
        bValidate_type = True
        irow = 0
        tbl.clear()
        tbl.setRowCount(0)
        for key, field in neccessary_fileds.items():
            tbl.insertRow(irow)

            newItem = QTableWidgetItem(key)
            newItem.setFlags(QtCore.Qt.ItemIsEnabled)
            tbl.setItem(irow, 0, newItem)

            combo = QComboBox()
            if field[1] == "num":
                combo.addItems(field_num_list)
            elif field[1] == "str":
                combo.addItems(field_str_list)

            field_index = layer.FindFieldIndex(field[0], False)
            if field_index > -1:
                if field[1] == "num":
                    if field[0] not in field_num_list:
                        bValidate_type = False
                elif field[1] == "str":
                    if field[0] not in field_str_list:
                        bValidate_type = False
                combo.setCurrentText(field[0])
            else:
                combo.setCurrentText("")
                bValidate_name = False

            tbl.setCellWidget(irow, 1, combo)

            irow += 1
        del datasource

        return bValidate_name, bValidate_type

    def add_spatial_data(self):
        fileName, fileType = QFileDialog.getOpenFileName(
            self, "选择待转换矢量图层文件", os.getcwd(),
            "ESRI Shapefile(*.shp)")

        if len(fileName) == 0:
            return False, "", None

        wks = workspaceFactory().get_factory(DataType.shapefile)
        datasource = wks.openFromFile(fileName)

        del wks

        if datasource is None:
            # layer = datasource.GetLayer(0)
            log.error("无法读取shp文件！{}".format(fileName), dialog=True)
            return False, "", None

        return True, fileName, datasource

    @Slot(QAbstractButton)
    def buttonBox_clicked(self, button: QAbstractButton):
        ds_Grid = None
        ds_PotentialLand = None

        try:
            if button == self.buttonBox.button(QDialogButtonBox.Ok):
                log.info("模型输入空间数据合规性检查...")

                path_Grid = self.txt_GridFile.text()
                path_PotentialLand = self.txt_PotentialLandFile.text()

                if path_Grid == "" or path_PotentialLand == "":
                    return

                wks = workspaceFactory().get_factory(DataType.shapefile)

                ds_Grid = wks.openFromFile(path_Grid)
                if ds_Grid is None:
                    raise IOError("读取shp文件{}发生错误！".format(path_Grid))

                ds_PotentialLand = wks.openFromFile(path_PotentialLand)
                if ds_PotentialLand is None:
                    raise IOError("读取shp文件{}发生错误！".format(path_PotentialLand))

                self.vGrid_field = self.check_field(ds_Grid, self.tbl_GridField, "标准单元")
                self.vPotential_field = self.check_field(ds_PotentialLand, self.tbl_PotentialLandField, "潜力用地")

                lyr_Grid = ds_Grid.GetLayer(0)
                lyr_PotentialLand = ds_PotentialLand.GetLayer(0)

                log.info("模型输入参数检查...")
                self.check_model_param()

                cur_path, filename = os.path.split(os.path.abspath(sys.argv[0]))
                res_path = os.path.join(cur_path, 'res')
                if not os.path.exists(res_path):
                    os.mkdir(res_path)

                modelCal(path_Grid, path_PotentialLand, lyr_Grid.GetName(), lyr_PotentialLand.GetName(),
                         self.vGrid_field, self.vPotential_field)

                print("OK")

            elif button == self.buttonBox.button(QDialogButtonBox.Cancel):
                self.threadTerminate()
                self.close()
        except IOError as err:
            log.error("模型无法运算. 原因:{}".format(err), dialog=True)
        except:
            log.error("模型无法运算. 原因:{}".format(traceback.format_exc()), dialog=True)
        finally:
            del ds_Grid
            del ds_PotentialLand

    #  检查空间数据字段合规性
    def check_field(self, ds, tbl, layer_name):
        vfields = {}

        for irow in range(tbl.rowCount()):
            k = tbl.item(irow, 0).text()
            v = tbl.cellWidget(irow, 1).currentText()

            if k == "" or v == "":
                raise IOError('{}图层缺失必要字段"{}"'.format(layer_name, k))

            layer = ds.GetLayer(0)
            field_index = layer.FindFieldIndex(v, False)
            if field_index < 0:
                raise IOError('{}图层不存在字段"{}"'.format(layer_name, v))
            # vfields[k] = field_index
            vfields[k] = v

        return vfields

    def validateValue(self):
        # doubleValidator = QDoubleValidator()
        reg = QRegularExpression(r"^(0[\.][0-9]{1,2})|1$")
        # doubleValidator.setNotation(QDoubleValidator.StandardNotation)
        doubleValidator = QRegularExpressionValidator()
        doubleValidator.setRegularExpression(reg)
        self.txt_type_gygjz.setValidator(doubleValidator)
        self.txt_type_zzgjz.setValidator(doubleValidator)
        self.txt_type_tdzb.setValidator(doubleValidator)
        self.txt_type_csgx.setValidator(doubleValidator)

        self.txt_indicator_ggfwsp.setValidator(doubleValidator)
        self.txt_indicator_jtkdx .setValidator(doubleValidator)
        self.txt_indicator_zzphzs.setValidator(doubleValidator)
        self.txt_indicator_ccjzl.setValidator(doubleValidator)
        self.txt_indicator_xzjjl.setValidator(doubleValidator)
        self.txt_indicator_xtzs.setValidator(doubleValidator)

    #  检查模型参数合规性
    def check_model_param(self):
        if self.txt_type_csgx == "":
            raise IOError("城市更新计划用地比例未设置！")
        if self.txt_type_tdzb == "":
            raise IOError("土地整备计划用地比例未设置！")
        if self.txt_type_zzgjz == "":
            raise IOError("旧住宅区改居住用地比例未设置！")
        if self.txt_type_gygjz == "":
            raise IOError("旧工业区改居住用地比例未设置！")

        if self.txt_indicator_xzjjl == "":
            raise IOError("新增建筑量权重未设置！")
        if self.txt_indicator_ccjzl == "":
            raise IOError("拆除建筑量权重未设置！")
        if self.txt_indicator_xtzs == "":
            raise IOError("用地形态指数权重未设置！")
        if self.txt_indicator_zzphzs == "":
            raise IOError("职住平衡指数权重未设置！")
        if self.txt_indicator_jtkdx == "":
            raise IOError("交通可达性权重未设置！")
        if self.txt_indicator_ggfwsp == "":
            raise IOError("公共服务水平权重未设置！")

        if float(self.txt_indicator_ggfwsp.text()) + float(self.txt_indicator_jtkdx.text()) + \
            float(self.txt_indicator_zzphzs.text()) + float(self.txt_indicator_xzjjl.text()) + \
            float(self.txt_indicator_xtzs.text()) + float(self.txt_indicator_ccjzl.text()) != 1:
            raise IOError("评价权重参数之和必须等于1，请重新调整！")

    def threadStop(self, bflag=True):
        self.thread.quit()
        if bflag:
            QMessageBox.information(self, "提示", "所有操作已完成，详细信息见日志列表.", QMessageBox.Close)
        else:
            QMessageBox.critical(self, "错误", "操作过程中发生异常，详细信息见日志列表.", QMessageBox.Close)

    def threadTerminate(self):
        try:
            if self.thread.isRunning():
                self.thread.terminate()
                self.thread.wait()
                del self.thread
            else:
                self.thread.quit()
                self.thread.wait()
        except:
            return


if __name__ == '__main__':
    ogr.UseExceptions()
    app = QApplication(sys.argv)
    # style = QStyleFactory.create("windows")
    # app.setStyle(style)
    window = frmModelCal()
    window.setWindowFlags(Qt.Window)
    window.show()
    sys.exit(app.exec_())